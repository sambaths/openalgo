from dataclasses import dataclass
from typing import List, Dict
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image
import math
import argparse
import yaml
from rich.console import Console
from rich.table import Table
from rich import box
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeRemainingColumn, TimeElapsedColumn
from rich.panel import Panel
import warnings
from pathlib import Path
import importlib
from fyers_utils import FyersBroker
from data_provider import get_data_provider
from cost_calculator import FeeConfig, TransactionCostCalculator
import inspect  # Add this import at the top with other imports

# Suppress warnings
pd.options.mode.chained_assignment = None
warnings.filterwarnings('ignore')

# Replace the current style configuration with:

# Configure plotting style
plt.style.use('default')  # Reset to default style
sns.set_style("darkgrid")  # Set seaborn style
sns.set_context("notebook", font_scale=1.2)  # Set context with larger font size
sns.set_palette("husl")  # Set color palette

# Update plot parameters
plt.rcParams.update({
    'figure.figsize': (12, 8),
    'axes.titlesize': 14,
    'axes.labelsize': 12,
    'xtick.labelsize': 10,
    'ytick.labelsize': 10,
    'lines.linewidth': 2,
    'grid.alpha': 0.3
})

# Constants
IST = pytz.timezone('Asia/Kolkata')
console = Console()

def convert_to_datetime(timestamp: int) -> datetime:
    return datetime.fromtimestamp(timestamp)

def convert_to_timestamp_with_tz(val: datetime) -> int:
    return datetime.fromtimestamp(int(val), tz=pytz.UTC).astimezone(IST)

def calculate_start_date(market_date_str: str, resolution: str, max_lookback_candles: int) -> str:
    """
    Calculate the start date for historical data based on required lookback candles.
    
    Args:
        market_date_str: Current market date in "YYYY-MM-DD" format
        resolution: Candle resolution in minutes
        max_lookback_candles: Maximum lookback period required in candles
        
    Returns:
        str: Computed start date in "YYYY-MM-DD" format
    """
    # Define market open and close times (IST)
    market_date = datetime.strptime(market_date_str, "%Y-%m-%d")
    
    # Define market open and close times
    market_open = market_date.replace(hour=9, minute=15, second=0, microsecond=0)
    market_close = market_date.replace(hour=15, minute=30, second=0, microsecond=0)
    
    # Convert times to IST
    market_open = market_open.replace(tzinfo=pytz.UTC).astimezone(IST)
    market_close = market_close.replace(tzinfo=pytz.UTC).astimezone(IST)
    
    # Calculate total trading seconds per day and candles per day
    trading_seconds = (market_close - market_open).total_seconds()
    candles_per_day = trading_seconds / (int(resolution) * 60)
    
    # Calculate required days
    days_required = math.ceil(max_lookback_candles / candles_per_day)
    
    # Add extra days for weekends and holidays (approximately 40% more days)
    days_required = math.ceil(days_required * 1.4)
    
    # Compute start date
    start_date = market_date - timedelta(days=days_required)
    
    return start_date.strftime("%Y-%m-%d")

@dataclass
class TradeMetrics:
    """Class to hold trade-specific metrics."""
    entry_price: float
    exit_price: float
    entry_time: datetime
    exit_time: datetime
    position_type: str
    position_size: float
    pnl: float
    roi: float
    duration: timedelta
    max_adverse_excursion: float
    max_favorable_excursion: float
    risk_reward_ratio: float
    trade_quality_score: float
    slippage_cost: float
    commision_cost: float
    net_pnl: float

@dataclass
class StrategyMetrics:
    """Class to hold strategy performance metrics."""
    total_trades: int
    winning_trades: int
    losing_trades: int
    win_rate: float
    profit_factor: float
    avg_profit_per_trade: float
    max_drawdown: float
    sharpe_ratio: float
    sortino_ratio: float
    calmar_ratio: float
    avg_trade_duration: timedelta
    max_consecutive_wins: int
    max_consecutive_losses: int
    value_at_risk: float
    expected_shortfall: float
    roi: float
    annualized_return: float
    volatility: float
    equity_curve: List[float]
    drawdown_curve: List[float]
    total_slippage_cost: float
    total_commision_cost: float
    cost_to_equity_ratio: float

class BacktestConfig:
    """Class to handle backtest configuration and validation."""
    
    def __init__(self, config_path: str):
        """
        Initialize backtest configuration.
        
        Args:
            config_path: Path to YAML configuration file
        """
        self.config = self._load_config(config_path)
        self._load_strategy_class()
        self._validate_config()
        
    def _load_config(self, config_path: str) -> dict:
        """Load and parse YAML configuration file."""
        try:
            with open(config_path, 'r') as file:
                return yaml.safe_load(file)
        except Exception as e:
            console.print(f"[red]Error loading configuration: {str(e)}[/red]")
            raise
            
    def _load_strategy_class(self):
        """Load strategy class from string specification."""
        try:
            if isinstance(self.config['strategy_class'], str):
                # Split the string into module path and class name
                module_path, class_name = self.config['strategy_class'].rsplit('.', 1)
                
                # Import the module
                module = importlib.import_module(module_path)
                
                # Get the class
                strategy_class = getattr(module, class_name)
                
                # Replace string with actual class
                self.config['strategy_class'] = strategy_class
                
        except Exception as e:
            console.print(f"[red]Error loading strategy class: {str(e)}[/red]")
            raise ValueError(f"Invalid strategy class specification: {self.config['strategy_class']}")
            
    def _validate_config(self):
        """Validate configuration parameters."""
        required_fields = ['tickers', 'start_date', 'end_date', 'strategy_class', 
                         'processor_params', 'risk_management']
        
        for field in required_fields:
            if field not in self.config:
                raise ValueError(f"Missing required configuration field: {field}")
                
        # Validate dates
        try:
            pd.to_datetime(self.config['start_date'])
            pd.to_datetime(self.config['end_date'])
        except:
            raise ValueError("Invalid date format in configuration")
            
        # Validate risk management parameters
        risk_params = self.config.get('risk_management', {})
        required_risk_params = ['max_position_size', 'risk_per_trade']
        
        for param in required_risk_params:
            if param not in risk_params:
                raise ValueError(f"Missing risk management parameter: {param}")

class TradeAnalyzer:
    """Class to analyze trade performance and calculate metrics."""
    
    def __init__(self, trades: List[dict], config: BacktestConfig):
        """
        Initialize trade analyzer.
        
        Args:
            trades: List of trade dictionaries
            config: Backtest configuration
        """
        self.trades = trades
        self.config = config
        self.metrics = self._calculate_metrics()
        
    def _calculate_metrics(self) -> StrategyMetrics:
        """Calculate comprehensive strategy metrics."""
        trades_df = pd.DataFrame(self.trades)
        
        if trades_df.empty:
            return self._get_empty_metrics()
        
        # Basic metrics
        total_trades = len(trades_df)
        winning_trades = len(trades_df[trades_df['net_pnl'] > 0])
        losing_trades = total_trades - winning_trades
        
        # Win rate and profit factor
        win_rate = winning_trades / total_trades if total_trades > 0 else 0
        total_profits = trades_df[trades_df['net_pnl'] > 0]['net_pnl'].sum()
        total_losses = abs(trades_df[trades_df['net_pnl'] < 0]['net_pnl'].sum())
        profit_factor = total_profits / total_losses if total_losses != 0 else float('inf')
        
        # Calculate trade durations
        trades_df['duration'] = pd.to_datetime(trades_df['exit_time'], utc=True) - pd.to_datetime(trades_df['entry_time'], utc=True)
        avg_duration = trades_df['duration'].mean()
        
        # Calculate consecutive wins/losses
        trades_df['win'] = trades_df['net_pnl'] > 0
        trades_df['loss'] = trades_df['net_pnl'] < 0
        consecutive_wins = self._get_max_consecutive(trades_df['win'])
        consecutive_losses = self._get_max_consecutive(trades_df['loss'])
        
        # Calculate returns and equity curve
        initial_capital = self.config.config.get('initial_capital', 100000)
        trades_df['returns'] = trades_df['net_pnl'] / initial_capital
        equity_curve = self._calculate_equity_curve(trades_df, initial_capital)
        
        # Risk metrics
        volatility = trades_df['returns'].std() * np.sqrt(252) if len(trades_df) > 1 else 0
        max_drawdown = self._calculate_max_drawdown(equity_curve)
        
        # Calculate risk-adjusted returns
        risk_free_rate = 0.02  # 2% annual risk-free rate
        excess_returns = trades_df['returns'] - (risk_free_rate / 252)
        
        sharpe = (np.mean(excess_returns) * np.sqrt(252)) / (np.std(excess_returns) if np.std(excess_returns) != 0 else float('inf'))
        
        # Sortino ratio calculation
        downside_returns = excess_returns[excess_returns < 0]
        sortino = (np.mean(excess_returns) * np.sqrt(252)) / (np.std(downside_returns) if len(downside_returns) > 0 and np.std(downside_returns) != 0 else float('inf'))
        
        # Calculate ROI and annualized return
        total_pnl = trades_df['net_pnl'].sum()
        roi = total_pnl / initial_capital
        
        # Calculate annualized return
        total_days = (trades_df['exit_time'].max() - trades_df['entry_time'].min()).days
        annualized_return = ((1 + roi) ** (365/total_days) - 1) if total_days > 0 else roi
        
        # Calmar ratio
        calmar = annualized_return / abs(max_drawdown) if max_drawdown != 0 else float('inf')
        
        # Value at Risk and Expected Shortfall
        var = np.percentile(trades_df['returns'], 5) if len(trades_df) > 0 else 0
        es = trades_df['returns'][trades_df['returns'] <= var].mean() if len(trades_df) > 0 else 0
        
        # Create drawdown curve
        drawdown_curve = self._calculate_drawdown_curve(equity_curve)
        
        # Calculate total commision and slippage costs
        total_commision_cost = trades_df['commision_cost'].sum() if 'commision_cost' in trades_df.columns else 0
        total_slippage_cost = trades_df['slippage_cost'].sum() if 'slippage_cost' in trades_df.columns else 0
        
        # Calculate cost to equity ratio
        cost_to_equity_ratio = (total_commision_cost + total_slippage_cost) / initial_capital if initial_capital > 0 else 0
        
        return StrategyMetrics(
            total_trades=total_trades,
            winning_trades=winning_trades,
            losing_trades=losing_trades,
            win_rate=max(0, min(1, win_rate)),  # Ensure between 0 and 1
            profit_factor=max(0, profit_factor),  # Ensure non-negative
            avg_profit_per_trade=trades_df['net_pnl'].mean(),
            max_drawdown=max(0, min(1, max_drawdown)),  # Ensure between 0 and 1
            sharpe_ratio=float(np.nan_to_num(sharpe, nan=0)),  # Handle NaN
            sortino_ratio=float(np.nan_to_num(sortino, nan=0)),  # Handle NaN
            calmar_ratio=float(np.nan_to_num(calmar, nan=0)),  # Handle NaN
            avg_trade_duration=avg_duration,
            max_consecutive_wins=consecutive_wins,
            max_consecutive_losses=consecutive_losses,
            value_at_risk=float(np.nan_to_num(var, nan=0)),  # Handle NaN
            expected_shortfall=float(np.nan_to_num(es, nan=0)),  # Handle NaN
            roi=roi,
            annualized_return=float(np.nan_to_num(annualized_return, nan=0)),  # Handle NaN
            volatility=float(np.nan_to_num(volatility, nan=0)),  # Handle NaN
            equity_curve=equity_curve,
            drawdown_curve=drawdown_curve,
            total_commision_cost=total_commision_cost,
            total_slippage_cost=total_slippage_cost,
            cost_to_equity_ratio=cost_to_equity_ratio
        )

    def _get_empty_metrics(self) -> StrategyMetrics:
        """Return empty metrics when no trades are available."""
        return StrategyMetrics(
            total_trades=0,
            winning_trades=0,
            losing_trades=0,
            win_rate=0.0,
            profit_factor=0.0,
            avg_profit_per_trade=0.0,
            max_drawdown=0.0,
            sharpe_ratio=0.0,
            sortino_ratio=0.0,
            calmar_ratio=0.0,
            avg_trade_duration=timedelta(0),
            max_consecutive_wins=0,
            max_consecutive_losses=0,
            value_at_risk=0.0,
            expected_shortfall=0.0,
            roi=0.0,
            annualized_return=0.0,
            volatility=0.0,
            equity_curve=[0.0],
            drawdown_curve=[0.0],
            total_commision_cost=0.0,
            total_slippage_cost=0.0,
            cost_to_equity_ratio=0.0
        )

    def _calculate_equity_curve(self, trades_df: pd.DataFrame, initial_capital: float) -> List[float]:
        """
        Calculate equity curve from trades using vectorized operations.
        
        Args:
            trades_df: DataFrame containing trade data
            initial_capital: Initial capital amount
            
        Returns:
            List of equity values over time
        """
        if trades_df.empty:
            return [initial_capital]
        
        # Ensure trades are sorted by entry time
        trades_df = trades_df.sort_values('entry_time')
        
        # Use vectorized operations for better performance
        cumulative_pnl = trades_df['net_pnl'].cumsum()
        equity_curve = initial_capital + cumulative_pnl
        
        # Return as list with initial capital as first value
        return [initial_capital] + equity_curve.tolist()

    def _calculate_drawdown_curve(self, equity_curve: List[float]) -> List[float]:
        """
        Calculate drawdown curve from equity curve using NumPy for vectorization.
        
        Args:
            equity_curve: List of equity values
            
        Returns:
            List of drawdown percentages
        """
        if not equity_curve or len(equity_curve) <= 1:
            return [0.0]
        
        # Convert to NumPy array for faster operations
        equity_array = np.array(equity_curve)
        
        # Calculate running maximum using vectorized operation
        running_max = np.maximum.accumulate(equity_array)
        
        # Calculate drawdowns as percentage of peak equity
        drawdowns = np.zeros_like(equity_array)
        mask = running_max > 0  # Avoid division by zero
        drawdowns[mask] = (equity_array[mask] - running_max[mask]) / running_max[mask]
        
        return drawdowns.tolist()

    def _get_max_consecutive(self, series: pd.Series) -> int:
        """
        Calculate maximum consecutive True values in series using optimized numpy operations.
        
        This implementation:
        1. Converts pandas Series to numpy array for faster operations
        2. Uses numpy's diff to detect changes (faster than pandas shift)
        3. Uses numpy's nonzero to find change points
        4. Calculates consecutive counts using pure numpy operations
        
        Args:
            series: pandas Series of boolean values
            
        Returns:
            int: Maximum number of consecutive True values
        """
        if series.empty:
            return 0
            
        # Convert to numpy array for faster operations
        arr = series.to_numpy()
        
        # Find indices where values change
        change_points = np.nonzero(np.diff(np.concatenate(([False], arr, [False]))))[0]
        
        # Calculate lengths of consecutive sequences
        lengths = change_points[1:] - change_points[:-1]
        
        # Filter lengths where the original values were True
        true_lengths = lengths[arr[change_points[:-1]]]
        
        # Return max length if any exist, otherwise 0
        return int(np.max(true_lengths)) if len(true_lengths) > 0 else 0

    def _calculate_max_drawdown(self, equity_curve: List[float]) -> float:
        """
        Calculate maximum drawdown from equity curve using NumPy for efficiency.
        
        Args:
            equity_curve: List of equity values
            
        Returns:
            Maximum drawdown as a positive percentage
        """
        if not equity_curve or len(equity_curve) <= 1:
            return 0.0
        
        # Convert to NumPy array
        equity_array = np.array(equity_curve)
        
        # Calculate running maximum
        running_max = np.maximum.accumulate(equity_array)
        
        # Calculate drawdowns
        drawdowns = (equity_array - running_max) / running_max
        
        # Return maximum drawdown as a positive percentage
        return abs(min(drawdowns, default=0.0))

class BacktestVisualizer:
    """Class to handle visualization of backtest results."""
    
    def __init__(self, results: Dict[str, StrategyMetrics], trades_data: Dict[str, List[dict]] = None, initial_capital: float = 100000):
        """
        Initialize visualizer with backtest results.
        
        Args:
            results: Dictionary mapping tickers to their metrics
            trades_data: Dictionary mapping tickers to their trade data
            initial_capital: Initial capital for portfolio calculations
        """
        self.results = results
        self.trades_data = trades_data
        self.initial_capital = initial_capital
        self.style_config = {
            'figure.figsize': (12, 8),
            'axes.titlesize': 14,
            'axes.labelsize': 12,
            'xtick.labelsize': 10,
            'ytick.labelsize': 10
        }
        plt.rcParams.update(self.style_config)

    def create_equity_curve_plot(self) -> str:
        """Create combined equity curve plot for all tickers."""
        plt.figure(figsize=(12, 6))
        
        for ticker, metrics in self.results.items():
            plt.plot(metrics.equity_curve, label=ticker)
            
        plt.title('Strategy Equity Curves')
        plt.xlabel('Trade Number')
        plt.ylabel('Equity')
        plt.legend()
        plt.grid(True)
        
        output_path = 'equity_curves.png'
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        return output_path

    def create_drawdown_plot(self) -> str:
        """Create drawdown visualization."""
        plt.figure(figsize=(12, 6))
        
        for ticker, metrics in self.results.items():
            plt.plot(metrics.drawdown_curve, label=ticker)
            
        plt.title('Drawdown Analysis')
        plt.xlabel('Trade Number')
        plt.ylabel('Drawdown %')
        plt.legend()
        plt.grid(True)
        
        output_path = 'drawdowns.png'
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        return output_path

    def create_performance_distribution(self) -> str:
        """Create performance distribution plot."""
        plt.figure(figsize=(12, 6))
        
        for ticker, metrics in self.results.items():
            sns.kdeplot(metrics.equity_curve, label=ticker)
            
        plt.title('Performance Distribution')
        plt.xlabel('Equity')
        plt.ylabel('Density')
        plt.legend()
        
        output_path = 'performance_dist.png'
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        return output_path
        
    def create_portfolio_equity_curve(self) -> str:
        """Create a combined portfolio equity curve across all tickers."""
        if not self.trades_data:
            return None
            
        # Combine all trades and sort by date
        all_trades = []
        for ticker, trades in self.trades_data.items():
            for trade in trades:
                if 'exit_time' in trade:  # Only include completed trades
                    all_trades.append({**trade, 'ticker': ticker})
        
        if not all_trades:
            return None
            
        # Sort trades by exit time
        all_trades = sorted(all_trades, key=lambda x: x['exit_time'])
        
        # Calculate cumulative P&L
        cumulative_pnl = [0]
        dates = [all_trades[0]['entry_time']]
        
        for trade in all_trades:
            cumulative_pnl.append(cumulative_pnl[-1] + trade['net_pnl'])
            dates.append(trade['exit_time'])
        
        # Create equity curve
        equity_curve = [self.initial_capital + pnl for pnl in cumulative_pnl]
        
        # Plot
        plt.figure(figsize=(12, 6))
        plt.plot(dates, equity_curve, 'b-', linewidth=2)
        plt.title('Portfolio Equity Curve (All Tickers Combined)')
        plt.xlabel('Date')
        plt.ylabel('Portfolio Value')
        plt.grid(True)
        
        # Format x-axis dates
        plt.gcf().autofmt_xdate()
        
        output_path = 'portfolio_equity.png'
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        return output_path
        
    def create_portfolio_drawdown(self) -> str:
        """Create a portfolio drawdown chart across all tickers."""
        if not self.trades_data:
            return None
            
        # Combine all trades and sort by date
        all_trades = []
        for ticker, trades in self.trades_data.items():
            for trade in trades:
                if 'exit_time' in trade:  # Only include completed trades
                    all_trades.append({**trade, 'ticker': ticker})
        
        if not all_trades:
            return None
            
        # Sort trades by exit time
        all_trades = sorted(all_trades, key=lambda x: x['exit_time'])
        
        # Calculate cumulative P&L
        cumulative_pnl = [0]
        dates = [all_trades[0]['entry_time']]
        
        for trade in all_trades:
            cumulative_pnl.append(cumulative_pnl[-1] + trade['net_pnl'])
            dates.append(trade['exit_time'])
        
        # Create equity curve
        equity_curve = np.array([self.initial_capital + pnl for pnl in cumulative_pnl])
        
        # Calculate drawdown
        running_max = np.maximum.accumulate(equity_curve)
        drawdown = (equity_curve - running_max) / running_max * 100  # Convert to percentage
        
        # Plot
        plt.figure(figsize=(12, 6))
        plt.fill_between(dates, drawdown, 0, color='red', alpha=0.3)
        plt.plot(dates, drawdown, 'r-', linewidth=1)
        plt.title('Portfolio Drawdown (All Tickers Combined)')
        plt.xlabel('Date')
        plt.ylabel('Drawdown %')
        plt.grid(True)
        
        # Format x-axis dates
        plt.gcf().autofmt_xdate()
        
        output_path = 'portfolio_drawdown.png'
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        return output_path
        
    def create_monthly_returns_heatmap(self) -> str:
        """Create a monthly returns heatmap."""
        if not self.trades_data:
            return None
            
        # Combine all trades
        all_trades = []
        for ticker, trades in self.trades_data.items():
            for trade in trades:
                if 'exit_time' in trade:  # Only include completed trades
                    all_trades.append({**trade, 'ticker': ticker})
        
        if not all_trades:
            return None
            
        # Create DataFrame with trade data
        trades_df = pd.DataFrame(all_trades)
        
        # Convert exit_time to datetime if it's not already, handling timezone-aware datetimes
        trades_df['exit_time'] = pd.to_datetime(trades_df['exit_time'], utc=True)
        
        # Extract year and month
        trades_df['year'] = trades_df['exit_time'].dt.year
        trades_df['month'] = trades_df['exit_time'].dt.month_name()
        trades_df['month'] = trades_df['month'].str.slice(0, 3)
        
        # Group by year and month, sum the P&L
        monthly_returns = trades_df.groupby(['year', 'month'])['net_pnl'].sum()
        
        # Convert to percentage of initial capital
        monthly_returns_pct = monthly_returns / self.initial_capital * 100
        
        # Create a pivot table for the heatmap
        pivot_table = monthly_returns_pct.unstack(level=0)
        
        # Plot heatmap
        plt.figure(figsize=(12, 8))
        ax = sns.heatmap(pivot_table, annot=True, fmt=".2f", cmap="RdYlGn", 
                         center=0, linewidths=.5, cbar_kws={'label': 'Return %'})
        
        # Set labels
        ax.set_title('Monthly Returns (%)')
        ax.set_xlabel('Year')
        ax.set_ylabel('Month')
        
        # Set month names instead of numbers
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month_names = [month for month in month_names if month in trades_df['month'].unique().tolist()] 
        
        ax.set_yticklabels(month_names, rotation=0)
        
        output_path = 'monthly_returns_heatmap.png'
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        return output_path

class BacktestReporter:
    """Class to generate detailed backtest reports."""
    
    def __init__(self, results: Dict[str, StrategyMetrics], config: BacktestConfig, trades_data: Dict[str, List[dict]], output_dir: str = "."):
        """
        Initialize reporter with backtest results and configuration.
        
        Args:
            results: Dictionary mapping tickers to their metrics
            config: Backtest configuration object
            trades_data: Dictionary mapping tickers to their trade data
            output_dir: Directory to save output files
        """
        self.results = results
        self.config = config
        self.trades_data = trades_data
        self.output_dir = Path(output_dir)
        initial_capital = self.config.config.get('initial_capital', 100000)
        self.visualizer = BacktestVisualizer(results, trades_data, initial_capital) if results else None
        
    def generate_cli_report(self):
        """Generate detailed CLI report using rich."""
        console.print("\n[bold cyan]═══ Backtest Results Summary ═══[/bold cyan]\n")
        
        # Strategy configuration table
        config_table = Table(title="Strategy Configuration", box=box.ROUNDED)
        config_table.add_column("Parameter", style="cyan")
        config_table.add_column("Value", style="green")
        
        # Add strategy init parameters
        strategy_class = self.config.config['strategy_class']
        
        # Get function signature
        signature = inspect.signature(strategy_class.__init__)
        
        # Extract parameters and their default values
        init_args = {param: value.default if value.default is not inspect.Parameter.empty else None
                    for param, value in signature.parameters.items() if param != 'self'}
        
        # Add a section for strategy init parameters
        config_table.add_row("Strategy Class", strategy_class.__name__)
        config_table.add_row("Strategy Parameters", "")
        
        for param_name, default_value in init_args.items():
            # Check if this parameter is provided in processor_params
            if param_name in self.config.config.get('processor_params', {}):
                actual_value = self.config.config['processor_params'][param_name]
                config_table.add_row(f"  - {param_name}", str(actual_value))
            else:
                config_table.add_row(f"  - {param_name}", f"(default: {default_value})")
        
        # Add other configuration parameters
        for key, value in self.config.config.items():
            if key != 'strategy_class' and key != 'processor_params':  # Skip these as we've handled them separately
                if isinstance(value, dict):
                    config_table.add_row(key, str({k: v for k, v in value.items()}))
                else:
                    config_table.add_row(key, str(value))
                
        console.print(config_table)
        console.print("\n")
        
        if not self.results:
            console.print("\n[red]No results to display. Check if data fetching was successful.[/red]")
            return
        
        # Performance metrics table
        metrics_table = Table(title="Performance Metrics by Ticker", box=box.ROUNDED)
        metrics_table.add_column("Metric", style="cyan")
        
        for ticker in self.results.keys():
            metrics_table.add_column(ticker, style="green")
            
        metrics_to_display = [
            ('Total Trades', 'total_trades'),
            ('Win Rate', 'win_rate', '{:.2%}'),
            ('Profit Factor', 'profit_factor', '{:.2f}'),
            ('Max Drawdown', 'max_drawdown', '{:.2%}'),
            ('Sharpe Ratio', 'sharpe_ratio', '{:.2f}'),
            ('Sortino Ratio', 'sortino_ratio', '{:.2f}'),
            ('Calmar Ratio', 'calmar_ratio', '{:.2f}'),
            ('ROI', 'roi', '{:.2%}'),
            ('Annualized Return', 'annualized_return', '{:.2%}'),
            ('Value at Risk (95%)', 'value_at_risk', '{:.2%}'),
            ('Expected Shortfall', 'expected_shortfall', '{:.2%}'),
            ('Max Consecutive Wins', 'max_consecutive_wins'),
            ('Max Consecutive Losses', 'max_consecutive_losses'),
            ('Total Commision Cost', 'total_commision_cost', '{:.2f}'),
            ('Total Slippage Cost', 'total_slippage_cost', '{:.2f}'),
            ('Cost to Equity Ratio', 'cost_to_equity_ratio', '{:.2%}')
        ]
        
        for metric_name, metric_attr, *format_spec in metrics_to_display:
            row_values = [metric_name]
            for ticker, metrics in self.results.items():
                value = getattr(metrics, metric_attr)
                if format_spec:
                    row_values.append(format_spec[0].format(value))
                else:
                    row_values.append(str(value))
            metrics_table.add_row(*row_values)
            
        console.print(metrics_table)

    def generate_excel_report(self):
        """Generate detailed Excel report with charts and metrics."""
        if not self.results:
            console.print("\n[red]No results to save. Excel report generation skipped.[/red]")
            return
        
        # Generate a filename with strategy name and timestamp
        strategy_name = self.config.config['strategy_class'].__name__
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.output_dir / f"{strategy_name}_backtest_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        self._create_summary_sheet(summary_sheet)
        
        # Detailed metrics sheet
        metrics_sheet = wb.create_sheet("Detailed Metrics")
        self._create_metrics_sheet(metrics_sheet)
        
        # Trades detail sheet for each ticker
        for ticker, trades in self.trades_data.items():
            trades_sheet = wb.create_sheet(f"{ticker} Trades")
            self._create_trades_sheet(trades_sheet, ticker, trades)
        
        # Daily activity sheet (timeline of trades)
        timeline_sheet = wb.create_sheet("Trade Timeline")
        self._create_timeline_sheet(timeline_sheet)
        
        # Raw data sheet for metric verification
        raw_data_sheet = wb.create_sheet("Raw Data")
        self._create_raw_data_sheet(raw_data_sheet)
        
        # Add charts
        charts_sheet = wb.create_sheet("Charts")
        self._add_charts(charts_sheet)
        
        wb.save(filename)
        console.print(f"\n[green]Excel report saved as {filename}[/green]")
        
    def _create_summary_sheet(self, sheet):
        """Create summary sheet with key metrics."""
        sheet.cell(1, 1, "Strategy Summary").font = Font(bold=True, size=14)
        
        # Add strategy configuration
        current_row = 3
        sheet.cell(current_row, 1, "Strategy Configuration").font = Font(bold=True)
        current_row += 1
        
        # Add strategy class info first
        strategy_class = self.config.config['strategy_class']
        sheet.cell(current_row, 1, "Strategy Class")
        sheet.cell(current_row, 2, strategy_class.__name__)
        current_row += 1
        
        # Add strategy init parameters section header
        current_row += 1
        sheet.cell(current_row, 1, "Strategy Parameters").font = Font(bold=True)
        current_row += 1
        
        # Get function signature and extract parameters
        signature = inspect.signature(strategy_class.__init__)
        init_args = {param: value.default if value.default is not inspect.Parameter.empty else None
                    for param, value in signature.parameters.items() if param != 'self'}
        
        # Add each parameter with its value
        for param_name, default_value in init_args.items():
            # Check if this parameter is provided in processor_params
            if param_name in self.config.config.get('processor_params', {}):
                actual_value = self.config.config['processor_params'][param_name]
                sheet.cell(current_row, 1, f"  - {param_name}")
                sheet.cell(current_row, 2, str(actual_value))
            else:
                sheet.cell(current_row, 1, f"  - {param_name}")
                sheet.cell(current_row, 2, f"(default: {default_value})")
            current_row += 1
        
        # Add other configuration parameters header
        current_row += 1
        sheet.cell(current_row, 1, "Other Configuration").font = Font(bold=True)
        current_row += 1
        
        # Add remaining configuration parameters
        for key, value in self.config.config.items():
            if key != 'strategy_class' and key != 'processor_params':  # Skip these as we've handled them separately
                sheet.cell(current_row, 1, key)
                if isinstance(value, dict):
                    sheet.cell(current_row, 2, str({k: v for k, v in value.items()}))
                else:
                    sheet.cell(current_row, 2, str(value))
                current_row += 1
        
        # Add key metrics summary
        current_row += 2
        sheet.cell(current_row, 1, "Key Metrics Summary").font = Font(bold=True)
        current_row += 1
        
        headers = ["Ticker", "Total Trades", "Win Rate", "Total Return", "Sharpe Ratio", "Max Drawdown"]
        for col, header in enumerate(headers, 1):
            sheet.cell(current_row, col, header).font = Font(bold=True)
            
        current_row += 1
        for ticker, metrics in self.results.items():
            sheet.cell(current_row, 1, ticker)
            sheet.cell(current_row, 2, metrics.total_trades)
            sheet.cell(current_row, 3, f"{metrics.win_rate:.2%}")
            sheet.cell(current_row, 4, f"{metrics.roi:.2%}")
            sheet.cell(current_row, 5, f"{metrics.sharpe_ratio:.2f}")
            sheet.cell(current_row, 6, f"{metrics.max_drawdown:.2%}")
            current_row += 1
    
    def _create_metrics_sheet(self, sheet):
        """Create detailed metrics sheet."""
        if not self.results:
            sheet.cell(1, 1, "No results available")
            return
        
        metrics_attributes = [attr for attr in dir(next(iter(self.results.values())))
                            if not attr.startswith('_') and 
                            not callable(getattr(next(iter(self.results.values())), attr))]
        
        # Write headers
        sheet.cell(1, 1, "Metric").font = Font(bold=True)
        for col, ticker in enumerate(self.results.keys(), 2):
            sheet.cell(1, col, ticker).font = Font(bold=True)
            
        # Write metrics
        for row, metric in enumerate(metrics_attributes, 2):
            sheet.cell(row, 1, metric)
            for col, (ticker, metrics) in enumerate(self.results.items(), 2):
                value = getattr(metrics, metric)
                if isinstance(value, (list, np.ndarray)):
                    continue
                if isinstance(value, float):
                    sheet.cell(row, col, value)
                else:
                    sheet.cell(row, col, str(value))
                    
    def _create_trades_sheet(self, sheet, ticker, trades):
        """Create trades sheet in Excel report."""
        # Handle no trades case
        if not trades:
            sheet.cell(1, 1, f"No trades for {ticker}")
            return
            
        # Set column headers
        headers = [
            'Entry Time', 'Exit Time', 'Duration', 'Position', 'Entry Price', 'Exit Price', 
            'Size', 'Gross P&L', 'Net P&L', 'ROI (%)', 'Stop Loss', 'Take Profit',
            'Transaction Costs', 'Brokerage', 'Regulatory Charges', 'Slippage', 'Trade Note'
        ]
        
        for col, header in enumerate(headers, 1):
            sheet.cell(1, col, header)
            sheet.cell(1, col).font = Font(bold=True)
            
        # Set column widths
        sheet.column_dimensions['A'].width = 18  # Entry Time
        sheet.column_dimensions['B'].width = 18  # Exit Time
        sheet.column_dimensions['C'].width = 12  # Duration
        sheet.column_dimensions['D'].width = 10  # Position
        sheet.column_dimensions['E'].width = 12  # Entry Price
        sheet.column_dimensions['F'].width = 12  # Exit Price
        sheet.column_dimensions['G'].width = 10  # Size
        sheet.column_dimensions['H'].width = 12  # Gross P&L
        sheet.column_dimensions['I'].width = 12  # Net P&L
        sheet.column_dimensions['J'].width = 10  # ROI
        sheet.column_dimensions['K'].width = 12  # Stop Loss
        sheet.column_dimensions['L'].width = 12  # Take Profit
        sheet.column_dimensions['M'].width = 15  # Transaction Costs
        sheet.column_dimensions['N'].width = 12  # Brokerage
        sheet.column_dimensions['O'].width = 18  # Regulatory Charges
        sheet.column_dimensions['P'].width = 12  # Slippage
        sheet.column_dimensions['Q'].width = 40  # Trade Note
        
        # Populate trade data
        for i, trade in enumerate(trades, 2):
            try:
                # Format times with timezone information
                entry_time = pd.to_datetime(trade['entry_time'], utc=True).astimezone(IST)
                exit_time = pd.to_datetime(trade['exit_time'], utc=True).astimezone(IST)
                
                # Calculate duration
                duration = exit_time - entry_time
                
                row = i
                sheet.cell(row, 1, entry_time.strftime('%Y-%m-%d %H:%M:%S'))
                sheet.cell(row, 2, exit_time.strftime('%Y-%m-%d %H:%M:%S'))
                sheet.cell(row, 3, str(duration))
                sheet.cell(row, 4, trade['position_type'])
                sheet.cell(row, 5, trade['entry_price'])
                sheet.cell(row, 6, trade['exit_price'])
                sheet.cell(row, 7, trade['position_size'])
                sheet.cell(row, 8, trade.get('gross_pnl', 0))
                sheet.cell(row, 9, trade.get('net_pnl', 0))
                roi = (trade.get('net_pnl', 0) / (trade['position_size'] * trade['entry_price'])) * 100 if trade['position_size'] > 0 else 0
                sheet.cell(row, 10, roi)
                sheet.cell(row, 11, trade.get('stop_loss', 0))
                sheet.cell(row, 12, 'Yes' if trade.get('take_profit', False) else 'No')
                
                # Add transaction cost details
                sheet.cell(row, 13, trade.get('total_transaction_costs', trade.get('commision_cost', 0) + trade.get('slippage_cost', 0)))
                sheet.cell(row, 14, trade.get('commision_cost', 0))
                sheet.cell(row, 15, trade.get('regulatory_costs', 0))
                sheet.cell(row, 16, trade.get('slippage_cost', 0))
                
                # Add trade note
                sheet.cell(row, 17, trade.get('note', ''))
                
                # Conditional formatting for P&L
                if trade.get('net_pnl', 0) > 0:
                    sheet.cell(row, 9).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                else:
                    sheet.cell(row, 9).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            except Exception as e:
                console.print(f"[red]Error processing trade {i-1}: {str(e)}[/red]")
                continue
    
    def _create_timeline_sheet(self, sheet):
        """Create a timeline view of trades across all tickers."""
        # Combine all trades from all tickers
        all_trades = []
        for ticker, trades in self.trades_data.items():
            for trade in trades:
                if 'exit_time' in trade:  # Only include completed trades
                    all_trades.append({**trade, 'ticker': ticker})
        
        if not all_trades:
            sheet.cell(1, 1, "No completed trades available")
            return
            
        # Sort trades by entry date
        all_trades = sorted(all_trades, key=lambda x: x['entry_time'])
        
        # Define headers
        headers = [
            "Date", "Ticker", "Action", "Price", "Position Type", 
            "Position Size", "P&L", "Cumulative P&L"
        ]
        
        # Set headers
        for col, header in enumerate(headers, 1):
            sheet.cell(1, col, header).font = Font(bold=True)
            
        # Track cumulative P&L
        cumulative_pnl = 0
        current_row = 2
        
        # Write timeline data - entries and exits as separate rows
        for trade in all_trades:
            # Entry row
            sheet.cell(current_row, 1, trade['entry_time'].strftime("%Y-%m-%d %H:%M:%S"))
            sheet.cell(current_row, 2, trade['ticker'])
            sheet.cell(current_row, 3, "ENTRY")
            sheet.cell(current_row, 4, trade['entry_price'])
            sheet.cell(current_row, 5, trade['position_type'])
            sheet.cell(current_row, 6, trade['position_size'])
            sheet.cell(current_row, 7, "")  # No P&L at entry
            sheet.cell(current_row, 8, cumulative_pnl)
            current_row += 1
            
            # Exit row
            if 'exit_time' in trade:
                sheet.cell(current_row, 1, trade['exit_time'].strftime("%Y-%m-%d %H:%M:%S"))
                sheet.cell(current_row, 2, trade['ticker'])
                sheet.cell(current_row, 3, "EXIT")
                sheet.cell(current_row, 4, trade['exit_price'])
                sheet.cell(current_row, 5, trade['position_type'])
                sheet.cell(current_row, 6, trade['position_size'])
                sheet.cell(current_row, 7, trade.get('net_pnl', 0))
                
                # Update and display cumulative P&L
                cumulative_pnl += trade.get('net_pnl', 0)
                sheet.cell(current_row, 8, cumulative_pnl)
                
                # Color code profitable exits
                if trade.get('net_pnl', 0) > 0:
                    for col in range(1, 9):
                        sheet.cell(current_row, col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif trade.get('net_pnl', 0) < 0:
                    for col in range(1, 9):
                        sheet.cell(current_row, col).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                current_row += 1
    
    def _create_raw_data_sheet(self, sheet):
        """Create a sheet with raw data for metric verification."""
        # Define headers based on the metrics we want to verify
        sheet.cell(1, 1, "Raw Data for Metric Verification").font = Font(bold=True, size=14)
        sheet.cell(3, 1, "This sheet contains raw data to help verify calculated metrics.").font = Font(italic=True)
        
        current_row = 5
        
        # Add trade statistics
        for ticker, trades in self.trades_data.items():
            sheet.cell(current_row, 1, f"Trade Statistics for {ticker}").font = Font(bold=True)
            current_row += 1
            
            if not trades:
                sheet.cell(current_row, 1, "No trades available")
                current_row += 2
                continue
                
            # Raw trade data table
            headers = ["Trade #", "Entry Date", "Exit Date", "Position", "Entry Price", 
                      "Exit Price", "Position Size", "P&L", "Return %", "Duration (days)"]
            
            for col, header in enumerate(headers, 1):
                sheet.cell(current_row, col, header).font = Font(bold=True)
            current_row += 1
            
            # Trade details for verification
            for i, trade in enumerate(trades, 1):
                if 'exit_time' not in trade:  # Skip incomplete trades
                    continue
                    
                duration_days = (trade['exit_time'] - trade['entry_time']).total_seconds() / (24 * 3600)
                return_pct = (trade.get('net_pnl', 0) / trade['position_size'] * trade['entry_price']) * 100 if trade['position_size'] > 0 else 0
                
                sheet.cell(current_row, 1, i)
                sheet.cell(current_row, 2, trade['entry_time'].strftime("%Y-%m-%d"))
                sheet.cell(current_row, 3, trade['exit_time'].strftime("%Y-%m-%d"))
                sheet.cell(current_row, 4, trade['position_type'])
                sheet.cell(current_row, 5, trade['entry_price'])
                sheet.cell(current_row, 6, trade['exit_price'])
                sheet.cell(current_row, 7, trade['position_size'])
                sheet.cell(current_row, 8, trade.get('net_pnl', 0))
                sheet.cell(current_row, 9, return_pct)
                sheet.cell(current_row, 10, duration_days)
                current_row += 1
            
            # Add intermediate data used in metrics calculations
            current_row += 2
            sheet.cell(current_row, 1, "Intermediate Calculations").font = Font(bold=True)
            current_row += 1
            
            # Win/loss statistics
            completed_trades = [t for t in trades if 'exit_time' in t]
            winning_trades = [t for t in completed_trades if t.get('net_pnl', 0) > 0]
            losing_trades = [t for t in completed_trades if t.get('net_pnl', 0) <= 0]
            
            win_rate = len(winning_trades) / len(completed_trades) if completed_trades else 0
            total_profit = sum(t.get('net_pnl', 0) for t in winning_trades)
            total_loss = abs(sum(t.get('net_pnl', 0) for t in losing_trades))
            profit_factor = total_profit / total_loss if total_loss != 0 else float('inf')
            
            # Display intermediate values
            metrics = [
                ("Total Trades", len(completed_trades)),
                ("Winning Trades", len(winning_trades)),
                ("Losing Trades", len(losing_trades)),
                ("Win Rate", win_rate),
                ("Total Profit", total_profit),
                ("Total Loss", total_loss),
                ("Profit Factor", profit_factor if profit_factor != float('inf') else "∞"),
                ("Initial Capital", self.config.config.get('initial_capital', 100000)),
                ("Final Capital", self.config.config.get('initial_capital', 100000) + sum(t.get('net_pnl', 0) for t in completed_trades))
            ]
            
            for i, (metric, value) in enumerate(metrics):
                sheet.cell(current_row + i, 1, metric)
                sheet.cell(current_row + i, 2, value)
                
            current_row += len(metrics) + 3  # Add space between tickers
        
    def _add_charts(self, sheet):
        """Add visualization charts to the Excel report."""
        # Add individual ticker charts
        img_path = self.visualizer.create_equity_curve_plot()
        img = Image(img_path)
        sheet.add_image(img, 'A1')
        
        img_path = self.visualizer.create_drawdown_plot()
        img = Image(img_path)
        sheet.add_image(img, 'A20')
        
        img_path = self.visualizer.create_performance_distribution()
        img = Image(img_path)
        sheet.add_image(img, 'A40')
        
        # Add portfolio-level charts
        portfolio_equity_path = self.visualizer.create_portfolio_equity_curve()
        if portfolio_equity_path:
            img = Image(portfolio_equity_path)
            sheet.add_image(img, 'M1')
        
        portfolio_drawdown_path = self.visualizer.create_portfolio_drawdown()
        if portfolio_drawdown_path:
            img = Image(portfolio_drawdown_path)
            sheet.add_image(img, 'M20')
        
        monthly_returns_path = self.visualizer.create_monthly_returns_heatmap()
        if monthly_returns_path:
            img = Image(monthly_returns_path)
            sheet.add_image(img, 'M40')


class Backtest:
    """Main class for running backtests with enhanced metrics and reporting."""
    
    def __init__(self, config: BacktestConfig, output_dir: str = "."):
        """
        Initialize backtest with configuration.
        
        Args:
            config: BacktestConfig object containing backtest parameters
            output_dir: Directory to save output files
        """
        self.config = config
        self.results = {}
        self.trades = {}
        self.reporter = None
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize data provider based on configuration
        data_provider_config = self.config.config.get('data_provider', {})
        provider_type = data_provider_config.get('type', 'fyers')
        provider_args = data_provider_config.get('args', {})
        
        self.data_provider = get_data_provider(provider_type, **provider_args)
        
        # Initialize transaction cost calculator
        self.cost_calculator = TransactionCostCalculator(
            FeeConfig.from_dict(self.config.config)
        )
        
        # Keep Fyers broker for margin calculation
        self.fyers_broker = FyersBroker()
        self.margin_dict = self.fyers_broker.get_margin([f"NSE:{ticker}-EQ" for ticker in self.config.config['tickers']])

    def run(self):
        """Execute backtest across all tickers."""
        console.print(Panel.fit(
            "[bold cyan]QuantForge Backtesting[/bold cyan]\n"
            f"Strategy: {self.config.config['strategy_class'].__name__}\n"
            f"Period: {self.config.config['start_date']} to {self.config.config['end_date']}\n"
            f"Resolution: {self.config.config['resolution']} minute bars\n"
            f"Initial Capital: {self.config.config['initial_capital']:,}\n"
        ))
        
        successful_tickers = 0
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            "[progress.percentage]{task.percentage:>3.1f}%",
            TimeElapsedColumn(),
            TimeRemainingColumn(elapsed_when_finished=True),
            console=console,
        ) as progress:
            
            task = progress.add_task("[cyan]Running backtest...", total=len(self.config.config['tickers']), transient=True)
            
            for i, ticker in enumerate(self.config.config['tickers']):
                progress.update(task, description=f"[cyan]Processing {ticker} ({i+1}/{len(self.config.config['tickers'])})")
                if self._backtest_ticker(ticker, progress):
                    successful_tickers += 1
                progress.update(task, advance=1)
                
        if successful_tickers == 0:
            console.print("[red]No successful backtests. Check data availability and configuration.[/red]")
            return
            
        self._generate_reports()
        
    def _backtest_ticker(self, ticker: str, progress: Progress) -> bool:
        """
        Run backtest for a single ticker.
        
        This method handles the core backtesting logic including:
        1. Data preparation and loading
        2. Strategy execution for each candle
        3. Trade entry, exit, and partial exit (take profit) handling
        4. Accurate transaction cost calculation for each trade action
        5. P&L calculation and tracking
        
        Take profit implementation details:
        - Take profit allows partial position closure when a profit target is hit
        - When a take profit signal is triggered, a portion of the position (defined by 
          position_to_close in the signal) is closed at the take_profit_price
        - All transaction costs (commissions, regulatory fees, slippage) are applied to the 
          partial exit, just as they would be for a full exit
        - The remaining position stays open until a full exit signal or another take profit
        - Capital is updated immediately upon each take profit action
        
        Args:
            ticker: Stock ticker symbol
            progress: Progress tracker for displaying progress
            
        Returns:
            bool: True if backtest was successful, False otherwise
        """
        all_trades = []
        all_data = []
        equity_curve = []
        drawdown_data = []
        capital = self.config.config.get('initial_capital', 100000)
        initial_capital = capital
        risk_params = self.config.config['risk_management']
        processor_params = self.config.config['processor_params'].copy()
        
        slippage_pct = risk_params.get('slippage', 0.001)
        commission_pct = risk_params.get('commission', 0.0025)
        
        # Create date range for backtesting
        start_date = pd.to_datetime(self.config.config['start_date'])
        end_date = pd.to_datetime(self.config.config['end_date'])
        trading_dates = pd.date_range(start=start_date, end=end_date, freq='B')  # 'B' for business days
        
        date_task_id = progress.add_task(f"[cyan]Processing {ticker} dates...", total=100, transient=True)
        
        total_gains = 0
        
        # Get main data for entire period using our data provider
        formatted_ticker = f"NSE:{ticker}-EQ"
        df_all = self.data_provider.get_historical_data(
            formatted_ticker,
            self.config.config.get('resolution', '5'),
            start_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d')
        )
        # Skip if no data available
        if df_all.empty:
            console.print(f"[yellow]No data available for {ticker}. Skipping.[/yellow]")
            progress.update(date_task_id, completed=True)
            return False
        
        # Structure data for processing
        df_all = self._structure_historical_data(ticker, df_all)

        # Process each trading day
        for current_date in trading_dates:
            
            # Update market date in processor params
            processor_params['market_date'] = current_date.strftime('%Y-%m-%d')
            df = df_all[df_all['datetime'].dt.date == current_date.date()]
            
            # Skip if empty dataframe
            if df.empty:
                progress.update(date_task_id, advance= 100 / len(trading_dates))
                continue

            # If strategy requires historical data, prepare it
            if processor_params.get('use_history', False):
                try:
                    # Calculate the needed date range
                    range_from = calculate_start_date(
                        current_date.strftime('%Y-%m-%d'),
                        processor_params.get('resolution', '5'),
                        processor_params.get('lookback_period', 20)
                    )
                    
                    range_to = (current_date - timedelta(days=1)).strftime('%Y-%m-%d')
                    
                    # First try to extract historical data from our main dataset
                    if not df_all.empty:
                        # Convert range_from to datetime for comparison
                        range_from_dt = pd.to_datetime(range_from)
                        range_to_dt = pd.to_datetime(range_to)
                        
                        # Filter df_all for the historical period we need
                        mask = (df_all['datetime'].dt.date >= range_from_dt.date()) & (df_all['datetime'].dt.date <= range_to_dt.date())
                        filtered_historical = df_all[mask]
                        
                        # Check if we have enough data
                        if len(filtered_historical) >= processor_params.get('lookback_period', 20):
                            processor_params['historical_data'] = filtered_historical
                        else:
                            # If not enough data in our main dataset, fetch it directly
                            # console.print(f"[cyan]Fetching additional historical data for {ticker} from {range_from} to {range_to}[/cyan]")
                            historical_df = self.data_provider.get_historical_data(
                                formatted_ticker,
                                processor_params.get('resolution', '5'),
                                range_from,
                                range_to
                            )
                            
                            if not historical_df.empty:
                                processor_params['historical_data'] = historical_df
                            else:
                                # console.print(f"[yellow]Could not retrieve historical data for {ticker}. Skipping date {current_date.strftime('%Y-%m-%d')}.[/yellow]")
                                progress.update(date_task_id, advance=100 / len(trading_dates))
                                continue
                
                except Exception as e:
                    console.print(f"[red]Error in historical data processing for {ticker} on {current_date.strftime('%Y-%m-%d')}: {str(e)}[/red]")
                    progress.update(date_task_id, advance=100 / len(trading_dates))
                    continue
            
            # Initialize strategy with updated parameters
            strategy = self.config.config['strategy_class'](**processor_params)
            
            # Ensure datetime column is properly formatted
            df['datetime'] = pd.to_datetime(df['datetime'])
            df['close'] = df['ltp']  # Add close column for compatibility
            df = df.sort_values('datetime').reset_index(drop=True)
            
            # Process each bar in the day
            entry_price = None
            entry_time = None
            position_size = 0
            position_closed_tp = 0
            realized_pnl_tp = 0
            take_profit_price = None
            take_profit_time = None
            daily_gains = 0
            for i, row in df.iterrows():
                # Generate signal using the run method
                output = strategy.run(row)
                
                # Add time to output
                output['time'] = row['datetime']
                
                # Store signal in dataframe
                if 'signal' in output:
                    df.loc[i, 'signal'] = output['signal']
                
                # Process trade actions
                if 'type' in output and output['type']:
                    try:
                        # Validate required fields in the output
                        if 'current_position' not in output:
                            # console.print(f"[yellow]Warning: Missing 'current_position' in output: {output}[/yellow]")
                            continue
                            
                        if output['type'] not in ['entry', 'exit', 'take_profit', None, 'in position']:
                            # console.print(f"[yellow]Warning: Unknown trade type: {output['type']}[/yellow]")
                            continue
                        
                        # Process trades based on position type and action type
                        if output['current_position'] == 'long':
                            if output['type'] == 'entry':
                                original_price = output['price']
                                entry_price = output['price'] * (1 + slippage_pct)
                                entry_time = output['time']
                                stop_loss = output['stop_loss']
                                
                                # Calculate position size based on risk
                                stop_loss_distance = output['price'] - output['stop_loss']
                                margin_multiplier = self.margin_dict.get(f"NSE:{ticker}-EQ", 1)
                                position_size = self._calculate_position_size(capital, output['price'], 
                                                                              risk_params['risk_per_trade'],
                                                                              stop_loss_pct=stop_loss_distance/output['price'],
                                                                              max_position_size=risk_params['max_position_size'])
                                position_size = position_size * margin_multiplier
                                
                                commision_cost = 0
                                # commision_cost = position_size * commission_pct * entry_price
                                slippage_cost = (entry_price - original_price) * position_size
                                # capital -= commision_cost
                                
                                position_closed_tp = 0
                                realized_pnl_tp = 0
                                realized_pnl = 0
                                take_profit_price = None
                                take_profit_time = None
                                
                            elif output['type'] == 'take_profit':
                                # Calculate take profit price with slippage for long positions (selling)
                                take_profit_original_price = output['price']
                                take_profit_price = output['price'] * (1 - slippage_pct)  # Apply slippage to sell price
                                position_to_close = output['position_to_close']
                                position_closed_tp = position_size * position_to_close
                                take_profit_time = output['time']
                                
                                # Calculate transaction costs for this partial exit
                                cost_result = self._calculate_transaction_costs(
                                    entry_price=entry_price,
                                    exit_price=take_profit_price,
                                    position_size=position_size,
                                    position_type='long',
                                    position_to_close=position_to_close
                                )
                                
                                # Extract results
                                realized_pnl_tp = cost_result['net_pnl']
                                tp_costs = cost_result['total_costs']
                                commision_cost += cost_result['commision_cost']
                                
                                # Calculate slippage cost separately (not included in transaction cost calculator)
                                tp_slippage_cost = (take_profit_price - take_profit_original_price) * position_closed_tp
                                slippage_cost += tp_slippage_cost
                                
                                # Update capital
                                capital += realized_pnl_tp
                                daily_gains += realized_pnl_tp
                                
                            elif output['type'] == 'exit':
                                # Calculate exit price with slippage for long positions (selling)
                                exit_original_price = output['price']
                                exit_price = output['price'] * (1 - slippage_pct)  # Apply slippage to sell price
                                position_to_close = output.get('position_to_close', 1.0)  # Default to closing full position
                                
                                # Calculate transaction costs for this exit
                                cost_result = self._calculate_transaction_costs(
                                    entry_price=entry_price,
                                    exit_price=exit_price,
                                    position_size=position_size,
                                    position_type='long',
                                    position_to_close=position_to_close
                                )
                                # print("###---", output['symbol'], output, cost_result)
                                # Extract results
                                gross_pnl = cost_result['gross_pnl']
                                net_pnl = cost_result['net_pnl']
                                exit_costs = cost_result['total_costs']
                                commision_cost += cost_result['commision_cost']
                                regulatory_costs = cost_result['regulatory_costs']
                                
                                # Calculate slippage cost separately (not included in transaction cost calculator)
                                exit_slippage_cost = (exit_original_price - exit_price) * position_size * position_to_close
                                slippage_cost += exit_slippage_cost
                                
                                # Create trade record using helper method
                                trade = self._create_trade_record(
                                    entry_price=entry_price,
                                    exit_price=output['price'],
                                    entry_time=entry_time,
                                    exit_time=output['time'],
                                    position_type=output['current_position'],
                                    position_size=position_size,
                                    net_pnl=net_pnl,
                                    gross_pnl=gross_pnl,
                                    realized_pnl_tp=realized_pnl_tp,
                                    stop_loss=stop_loss,
                                    take_profit_price=take_profit_price,
                                    capital=capital,
                                    commision_cost=commision_cost,
                                    regulatory_costs=regulatory_costs,
                                    slippage_cost=slippage_cost,
                                    total_costs=exit_costs,
                                    ticker=ticker,
                                    trade_id=output.get('trade_id', ''),
                                    position_closed_tp=position_closed_tp,
                                    take_profit_time=take_profit_time,
                                    note=output.get('note', None)
                                )
                                all_trades.append(trade)
                                capital += net_pnl 
                                daily_gains += net_pnl
                            
                            elif output['type'] == 'in position':
                                pass
                                
                                
                        elif output['current_position'] == 'short':
                            if output['type'] == 'entry':
                                original_price = output['price']
                                entry_price = output['price'] * (1 - slippage_pct)
                                entry_time = output['time']
                                stop_loss = output['stop_loss']
                                # Calculate position size based on risk
                                stop_loss_distance = output['stop_loss'] - output['price']
                                margin_multiplier = self.margin_dict.get(f"NSE:{ticker}-EQ", 1)
                                position_size = self._calculate_position_size(capital, output['price'], 
                                                    risk_params['risk_per_trade'],
                                                    stop_loss_pct=stop_loss_distance/output['price'],
                                                    max_position_size=risk_params['max_position_size'])
                                position_size = position_size * margin_multiplier
                                
                                commision_cost = 0
                                # commision_cost = position_size * commission_pct * entry_price
                                slippage_cost = (original_price - entry_price) * position_size
                                # capital -= commision_cost
                                
                                realized_pnl = 0
                                realized_pnl_tp = 0
                                position_closed_tp = 0
                                take_profit_price = None
                                take_profit_time = None
                                
                            elif output['type'] == 'take_profit':
                                # Calculate take profit price with slippage for short positions (buying)
                                take_profit_original_price = output['price']
                                take_profit_price = output['price'] * (1 + slippage_pct)  # Apply slippage to buy price
                                position_to_close = output['position_to_close']
                                position_closed_tp = position_size * position_to_close
                                take_profit_time = output['time']
                                
                                # Calculate transaction costs for this partial exit
                                cost_result = self._calculate_transaction_costs(
                                    entry_price=entry_price,
                                    exit_price=take_profit_price,
                                    position_size=position_size,
                                    position_type='short',
                                    position_to_close=position_to_close
                                )
                                
                                # Extract results
                                realized_pnl_tp = cost_result['net_pnl']
                                tp_costs = cost_result['total_costs']
                                commision_cost += cost_result['commision_cost']
                                
                                # Calculate slippage cost separately (not included in transaction cost calculator)
                                tp_slippage_cost = (take_profit_price - take_profit_original_price) * position_closed_tp
                                slippage_cost += tp_slippage_cost
                                
                                # Update capital
                                capital += realized_pnl_tp
                                daily_gains += realized_pnl_tp
                                
                            elif output['type'] == 'exit':
                                # Calculate exit price with slippage for short positions (buying)
                                exit_original_price = output['price']
                                exit_price = output['price'] * (1 + slippage_pct)  # Apply slippage to buy price
                                position_to_close = output.get('position_to_close', 1.0)  # Default to closing full position
                                
                                # Calculate transaction costs for this exit
                                cost_result = self._calculate_transaction_costs(
                                    entry_price=entry_price,
                                    exit_price=exit_price,
                                    position_size=position_size,
                                    position_type='short',
                                    position_to_close=position_to_close
                                )
                                
                                # Extract results
                                gross_pnl = cost_result['gross_pnl']
                                net_pnl = cost_result['net_pnl']
                                exit_costs = cost_result['total_costs']
                                commision_cost += cost_result['commision_cost']
                                regulatory_costs = cost_result['regulatory_costs']
                                
                                # Calculate slippage cost separately (not included in transaction cost calculator)
                                exit_slippage_cost = (exit_price - exit_original_price) * position_size * position_to_close
                                slippage_cost += exit_slippage_cost
                                
                                # Create trade record using helper method
                                trade = self._create_trade_record(
                                    entry_price=entry_price,
                                    exit_price=output['price'],
                                    entry_time=entry_time,
                                    exit_time=output['time'],
                                    position_type=output['current_position'],
                                    position_size=position_size,
                                    net_pnl=net_pnl,
                                    gross_pnl=gross_pnl,
                                    realized_pnl_tp=realized_pnl_tp,
                                    stop_loss=stop_loss,
                                    take_profit_price=take_profit_price,
                                    capital=capital,
                                    commision_cost=commision_cost,
                                    regulatory_costs=regulatory_costs,
                                    slippage_cost=slippage_cost,
                                    total_costs=exit_costs,
                                    ticker=ticker,
                                    trade_id=output.get('trade_id', ''),
                                    position_closed_tp=position_closed_tp,
                                    take_profit_time=take_profit_time,
                                    note=output.get('note', None)
                                )
                                all_trades.append(trade)
                                capital +=  net_pnl 
                                daily_gains += net_pnl
                        
                        else:
                            pass
                            # console.print(f"[yellow]Warning: Unknown position type: {output['current_position']}[/yellow]")
                    
                    except Exception as e:
                        # console.print(f"[red]Error processing trade action: {str(e)}[/red]")
                        import traceback
                        console.print(traceback.format_exc())
            
            # Update tracking variables
            total_gains += daily_gains
            all_data.append(df)
            equity_curve.append(capital)
            drawdown_data.append(capital)
            progress.update(date_task_id, advance= 100 / len(trading_dates))
        
        progress.remove_task(date_task_id)
        # No trades were made
        if not all_trades:
            console.print(f"[yellow]No trades generated for {ticker}[/yellow]")
            return False
        
        # Store full dataframe for analysis if needed (not affecting existing code)
        self.full_data = getattr(self, 'full_data', {})
        self.full_data[ticker] = pd.concat(all_data).reset_index(drop=True) if all_data else pd.DataFrame()
        
        # Store equity curve and drawdown data (not affecting existing code)
        self.equity_curves = getattr(self, 'equity_curves', {})
        self.equity_curves[ticker] = equity_curve
        
        self.drawdown_data = getattr(self, 'drawdown_data', {})
        self.drawdown_data[ticker] = drawdown_data
        
        self.total_gains = getattr(self, 'total_gains', {})
        self.total_gains[ticker] = total_gains
        
        # Analyze results using the existing TradeAnalyzer for compatibility
        analyzer = TradeAnalyzer(all_trades, self.config)
        self.results[ticker] = analyzer.metrics
        self.trades[ticker] = all_trades
        return True

    def _structure_historical_data(self, ticker: str, historical_data: dict) -> pd.DataFrame:
        """
        Structure historical data for backtesting.
        
        Args:
            ticker: Stock ticker symbol
            historical_data: Historical data in dict or DataFrame format
            
        Returns:
            pd.DataFrame: Structured DataFrame with all required fields
        """
        # Handle case when input is already a DataFrame
        if isinstance(historical_data, pd.DataFrame):
            return historical_data
            
        # Handle empty or invalid responses
        if not isinstance(historical_data, dict) or 'candles' not in historical_data or not historical_data['candles']:
            console.print(f"[yellow]Warning: Empty or invalid data structure for {ticker}[/yellow]")
            return pd.DataFrame()  # Return empty DataFrame
            
        # Create DataFrame from fetched data
        df = pd.DataFrame(historical_data['candles'], columns=['t', 'o', 'h', 'l', 'c', 'v'])
        
        # Convert timestamp to datetime with proper timezone
        df['datetime'] = pd.to_datetime(df['t'], unit='s').dt.tz_localize(pytz.UTC).dt.tz_convert(IST)
        
        # Calculate cumulative volume
        df['cum_vol'] = df['v'].cumsum()
        # Calculate previous close price using shift
        df['prev_close'] = df['c'].shift(1).fillna(df['o'].iloc[0])

        # Calculate change and change percentage
        df['ch'] = df['c'] - df['prev_close']
        df['chp'] = (df['ch'] / df['prev_close'] * 100).fillna(0)

        # Calculate average trade price
        df['avg_trade_price'] = (df['o'] + df['h'] + df['l'] + df['c']) / 4

        # Prepare the final DataFrame with required metrics
        metrics_df = pd.DataFrame({
            'datetime': df['datetime'],  # Keep the datetime column as is
            'symbol': f"NSE:{ticker}-EQ",
            'ltp': df['c'].round(2),
            'vol_traded_today': df['cum_vol'].astype(int),
            'last_traded_time': df['t'].astype(int),
            'exch_feed_time': df['t'].astype(int),
            'bid_size': 0,
            'ask_size': 0,
            'bid_price': (df['c'] - 0.05).round(2),
            'ask_price': (df['c'] + 0.05).round(2),
            'last_traded_qty': df['v'].astype(int),
            'tot_buy_qty': (df['cum_vol'] // 2).astype(int),
            'tot_sell_qty': (df['cum_vol'] - (df['cum_vol'] // 2)).astype(int),
            'avg_trade_price': df['avg_trade_price'].round(2),
            'low_price': df['l'].round(2),
            'high_price': df['h'].round(2),
            'open_price': df['o'].round(2),
            'prev_close_price': df['prev_close'].round(2),
            'type': "historical",
            'ch': df['ch'].round(2),
            'chp': df['chp'].round(2)
        })
        
        return metrics_df

    def _fetch_historical_data(self, ticker: str) -> pd.DataFrame:
        """
        Fetch and prepare historical data for backtesting.
        
        Args:
            ticker: Stock ticker symbol
            
        Returns:
            pd.DataFrame: Structured historical data
        """
        formatted_ticker = f"NSE:{ticker}-EQ"
        
        try:
            console.print(f"[cyan]Fetching historical data for {ticker}...[/cyan]")
            
            # Use the data provider to fetch historical data
            df = self.data_provider.get_historical_data(
                formatted_ticker,
                self.config.config.get('resolution', '5'),
                self.config.config.get('start_date'),
                self.config.config.get('end_date')
            )
            
            if df.empty:
                console.print(f"[yellow]No historical data found for {ticker}[/yellow]")
                return pd.DataFrame()
                
            return df
            
        except Exception as e:
            console.print(f"[red]Error fetching historical data for {ticker}: {str(e)}[/red]")
            return pd.DataFrame()
              
    def _calculate_position_size(self, capital: float, price: float, risk_per_trade:float, 
                               stop_loss_pct: float, max_position_size: float) -> float:
        """Calculate position size based on risk parameters."""
        risk_amount = capital * risk_per_trade
        price_risk = price * stop_loss_pct
        # print(risk_amount, price_risk, risk_amount / price_risk, capital * max_position_size, capital * max_position_size / price)
        position_size = min(risk_amount / price_risk, capital * max_position_size / price)
        return int(position_size)
        
    def _generate_reports(self):
        """Generate backtest reports."""
        if not self.results:
            console.print("[red]No results to report. Check if data fetching was successful for any symbols.[/red]")
            return
        
        self.reporter = BacktestReporter(self.results, self.config, self.trades, self.output_dir)
        self.reporter.generate_cli_report()
        self.reporter.generate_excel_report()

    def _calculate_transaction_costs(self, entry_price: float, exit_price: float, position_size: float, 
                                    position_type: str, position_to_close: float = 1.0):
        """
        Calculate transaction costs for a trade (entry and exit).
        
        Args:
            entry_price: Price at which the position was entered
            exit_price: Price at which the position is being exited
            position_size: Size of the position
            position_type: Type of position ('long' or 'short')
            position_to_close: Fraction of position being closed (default: 1.0 for full close)
            
        Returns:
            dict: Dictionary containing transaction costs and P&L information
        """
        # Calculate the quantity that is being closed
        quantity = int(position_size * position_to_close)
        if quantity <= 0:
            return {
                'gross_pnl': 0,
                'net_pnl': 0,
                'commision_cost': 0,
                'regulatory_costs': 0,
                'slippage_cost': 0,
                'total_costs': 0
            }
            
        # For long positions: buy at entry, sell at exit
        # For short positions: sell at entry, buy at exit
        if position_type == 'long':
            buy_price = entry_price
            sell_price = exit_price
            gross_pnl = (exit_price - entry_price) * quantity
        else:  # short
            buy_price = exit_price
            sell_price = entry_price
            gross_pnl = (entry_price - exit_price) * quantity
        
        # Calculate transaction costs using the cost calculator
        transaction_costs = self.cost_calculator.calculate_costs(
            buy_price=buy_price,
            sell_price=sell_price,
            quantity=quantity
        )
        
        # Extract cost components
        total_costs = transaction_costs['totals']['total_charges']
        commision_cost = transaction_costs['brokerage']['total_brokerage']
        regulatory_costs = transaction_costs['regulatory_charges']['total_regulatory_charges']
        
        # Calculate net P&L
        net_pnl = gross_pnl - total_costs
        
        return {
            'gross_pnl': gross_pnl,
            'net_pnl': net_pnl,
            'commision_cost': commision_cost,
            'regulatory_costs': regulatory_costs,
            'total_costs': total_costs,
            'transaction_details': transaction_costs  # Include full details for reference if needed
        }

    def _create_trade_record(self, entry_price, exit_price, entry_time, exit_time, 
                             position_type, position_size, net_pnl, gross_pnl, realized_pnl_tp, 
                             stop_loss, take_profit_price, capital, commision_cost, regulatory_costs, 
                             slippage_cost, total_costs, ticker, trade_id='', position_closed_tp=0, 
                             take_profit_time=None, note=None):
        """
        Create a standardized trade record.
        
        Args:
            entry_price: Price at entry
            exit_price: Price at exit
            entry_time: Time of entry
            exit_time: Time of exit
            position_type: Type of position ('long' or 'short')
            position_size: Size of position
            net_pnl: Net profit and loss
            gross_pnl: Gross profit and loss
            realized_pnl_tp: Realized P&L from take profit
            stop_loss: Stop loss price
            take_profit_price: Take profit price (None if not used)
            capital: Capital after trade
            commision_cost: Commission costs
            regulatory_costs: Regulatory charges
            slippage_cost: Slippage costs
            total_costs: Total transaction costs
            ticker: Stock ticker symbol
            trade_id: Unique trade ID
            position_closed_tp: Position size closed in take profit
            take_profit_time: Time of take profit
            note: Additional information about the trade (e.g., exit reason)
            
        Returns:
            dict: Trade record
        """
        # Calculate ROI safely
        roi = 0
        if entry_price and position_size and position_size > 0:
            roi = (net_pnl + realized_pnl_tp) / (entry_price * position_size)
            
        return {
            'entry_price': entry_price,
            'exit_price': exit_price,
            'entry_time': entry_time,
            'exit_time': exit_time,
            'position_type': position_type,
            'position_size': position_size,
            'gross_pnl': gross_pnl + realized_pnl_tp,
            'net_pnl': net_pnl + realized_pnl_tp,
            'roi': roi,
            'stop_loss': stop_loss,
            'take_profit': take_profit_price is not None,  # If take_profit_price exists, take profit was used
            'commision_cost': commision_cost,
            'regulatory_costs': regulatory_costs,
            'slippage_cost': slippage_cost,
            'total_transaction_costs': total_costs,
            
            # Additional fields for new approach
            'ticker': ticker,
            'trade_id': trade_id,
            'realized_pnl_tp': realized_pnl_tp,
            'capital': capital + net_pnl,
            'take_profit_price': take_profit_price,
            'position_closed_tp': position_closed_tp,
            'take_profit_time': take_profit_time,
            'note': note,  # Add the note field
        }

def main():
    """Command-line interface for the backtester."""
    parser = argparse.ArgumentParser(
        description="Quantitative Trading Strategy Backtester",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    
    parser.add_argument(
        '--config',
        type=str,
        required=True,
        help='Path to YAML configuration file'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        default='backtest_results',
        help='Output directory for results'
    )
    
    args = parser.parse_args()
    
    try:
        # Create output directory
        output_dir = Path(args.output)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Load configuration
        config = BacktestConfig(args.config)
        # df = pd.read_csv('https://raw.githubusercontent.com/sambaths/indices-data/refs/heads/main/data/index_constituents.csv')
        
        # df = df[(df['Market Cap Classification'] == 'Small Cap') &
        #         (df['Stock'] != df['Index']) &
        #         (df['series'] == 'EQ')].drop_duplicates(subset='Stock').sort_values(by='totalTradedVolume', ascending = False)
        # df = pd.read_csv('data/screen_candidates_smallcap.csv')
        # config.config['tickers'] = df['CSV Symbol'].unique().tolist()[0:5]
        
        # Run backtest
        backtest = Backtest(config, output_dir)
        backtest.run()
        
    except Exception as e:
        console.print(f"[red]Error during backtest execution: {str(e)}[/red]")
        raise

if __name__ == "__main__":
    main()
    
# python -m system.backtester.backtest --config system/backtester/backtest_config.yaml --output backtest_results